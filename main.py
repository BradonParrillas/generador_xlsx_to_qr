import openpyxl
import qrcode

# Excel
path_data = 'Data.xlsx'
path_qr_images = 'images\\'
workbook = openpyxl.load_workbook(path_data)
sheet_data = workbook.active
last_row = sheet_data.max_row

# Colors
fill_color_rgb = (37, 38, 94)
back_color_rgb = (255, 255, 255)

def generar_qr(data, name, description = ''):
    # QR
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=50,
        border=1,
    )

    qr.add_data(data)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color=fill_color_rgb, back_color=back_color_rgb)

    image_file = open(name, 'wb')
    qr_img.save(image_file)
    image_file.close()


for i in range(2, last_row + 1):
    title = sheet_data.cell(row=i, column=1).value
    qr_data = sheet_data.cell(row=i, column=2).value
    description = sheet_data.cell(row=i, column=3).value

    qr_name = path_qr_images + title[:-4] + ' - ' + description + '.png'

    generar_qr(qr_data, qr_name, description)
    print("QR generado: ", title[:-4], description)